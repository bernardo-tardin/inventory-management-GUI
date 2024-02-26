import PySimpleGUI as sg
import modulo1 as md
import io
import os
from os import listdir
from os.path import isfile,join
from datetime import datetime
from PIL import Image
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill,Font,Alignment,Border,Side

def alterarProduto(bd,produto):
    sg.theme('Reddit')
    tipos=[('JPG (*.jpg)', '*.jpg'),('PNG (*.png)','*.png')]
    
    interface=[
        [sg.Text('Nome do produto:', font='Helvetica 10'),sg.Input(produto[0], font='Helvetica 13',size=(30),key='-nome')],
        [sg.Text('Preço:', font='Helvetica 10'), sg.Input(produto[2],font='Helvetica 13',size=(8),key='-preco')],
        [sg.Text('Imagem do produto:', font='Helvetica 10'), sg.FileBrowse(produto[3],file_types=tipos,font='Helvetica 10', key='-image')],
        [sg.Button('Alterar', font='Helvetica 12', key='-alterar')]
        ]
    window = sg.Window(title='Alterar', font='Helvetica 24').Layout(interface)
    stop = False 
    while not stop:
        event, values = window.read() 
        if event == sg.WIN_CLOSED: 
            stop = True
        elif event == '-alterar':
            nome=values['-nome']
            preço=values['-preco']
            if values['-image']=='':
                image=produto[3]
            else:
                image=values['-image']
            new=md.adicionarProduto(bd,nome,preço,image)
            window.close()
            return new
    window.close()

def janelaAdd(bd):
    sg.theme('Reddit')
    tipos=[('JPG (*.jpg)', '*.jpg'),('PNG (*.png)','*.png')]
    
    interface=[
        [sg.Text('Nome do produto:', font='Helvetica 10'),sg.Input(font='Helvetica 13',size=(30),key='-nome')],
        [sg.Text('Preço:', font='Helvetica 10'), sg.Input(font='Helvetica 13',size=(8),key='-preco')],
        [sg.Text('Imagem do produto:', font='Helvetica 10'), sg.FileBrowse(file_types=tipos,font='Helvetica 10', key='-image')],
        [sg.Button('Adicionar Produto', font='Helvetica 12', key='-adicionar')]
        ]

    window = sg.Window(title='Adicionar', font='Helvetica 24').Layout(interface)
    stop = False 
    while not stop:
        event, values = window.read() 
        if event == sg.WIN_CLOSED: 
            stop = True
        elif event == '-adicionar':
            if values['-nome']=='' and values['-preco']=='':
                janelaErro('Insira o nome e o preço do produto')
            elif values['-preco']=='':
                janelaErro('Insira o preço do produto')
            elif values['-nome']=='':
                janelaErro('Insira o nome do produto')
            else:
                nome=values['-nome']
                preço=values['-preco']
                image=values['-image']
                novo=md.adicionarProduto(bd,nome,preço,image)
                window.close()
                return novo
    window.close()

def janelaErro(mensagem):
    interface = [
        [sg.Text(mensagem)],
        [sg.Button('Ok',font='Haveltica 12')]
    ]
    
    window = sg.Window(title='ERRO', font='Helvetica 14').Layout(interface)
    
    stop=False
    while not stop:
        event, values = window.read()
        if event == 'Ok' or event == sg.WIN_CLOSED:
            stop=True
    window.close()

def consultarProduto(bd):
    product=[]
    for produto in bd:
        product.append(produto[0])
    product.sort(key=lambda x:x[0].upper())
    
    coluna=[
        [sg.Text('Nome do produto:',font=('Heveltica',10,'bold')),sg.Text(key='-nome')],
        [sg.Text('Preço:',font=('Heveltica',10,'bold')),sg.Text(key='-preço')],
        [sg.Text('Imagem do produto:',font=('Heveltica',10,'bold'))],
        [sg.Image(key='-image',size=(5,10))]
    ]
    
    interface=[
        [sg.Text('Produtos', font='Helvetica 14')],
        [sg.Button('Consultar'),sg.Button('Remover'), sg.Button('Alterar')],
        [sg.Listbox(values=product,size=(18,10), key='-produtos'), sg.VerticalSeparator(), sg.Column(coluna)]
        ]
    
    
    window=sg.Window('Consultar',size=(500,400)).Layout(interface)
    
    stop=False
    while not stop:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            stop=True
        elif event == 'Consultar':
            if len(values['-produtos']) > 0:
                name=values['-produtos'][0]
                for produto in bd:
                    if produto[0] == name:
                        preco=produto[2]
                        if (len(produto))>=4:
                            fimagem=produto[3]
                            if os.path.exists(fimagem):
                                md.thumb(fimagem)
                                image=Image.open(fimagem)
                                bio=io.BytesIO()
                                image.save(bio,format='PNG')
                                window['-image'].update(data=bio.getvalue())
                        else:
                            image=''
                            window['-image'].update(image)
                window.find_element('-nome').update(name)
                window.find_element('-preço').update(preco)
            else:
                janelaErro('Selecione um produto antes de consultar')
        elif event == 'Remover':
            if len(values['-produtos']) > 0:
                product.remove(values['-produtos'][0])
                name=values['-produtos'][0]
                for p in bd:
                    if p[0]==name:
                        bd.remove(p)
                window.find_element('-produtos').update(product)
            else:
                janelaErro('Selecione um produto antes de remover')
        elif event == 'Alterar':
            if len(values['-produtos']) > 0:
                name=values['-produtos'][0]
                for produtos in bd:
                    if produtos[0] == name:
                        pro=produtos
                new=alterarProduto(BD,pro)
                if new != None:
                    bd.remove(pro)
                    bd.append(new)
                    product.remove(pro[0])
                    product.append(new[0])
                    product.sort(key=lambda x:x[0].upper())
                    window.find_element('-produtos').update(product)
            else:
                janelaErro('Selecione um produto antes de alterar')
    window.close()
    return bd

def criarNota(bd):
    sg.theme('Reddit')
    produtos=[]
    nota=[]
    bd2=[]
    for produto in bd:
        produtos.append(produto[0])
    produtos.sort(key=lambda x:x[0].upper())

    coluna1=[
        [sg.Text('Nome do produto:',font=('Heveltica',10,'bold')),sg.Text(key='-nome')],
        [sg.Text('Preço:',font=('Heveltica',10,'bold')),sg.Text(key='-preço')],
        [sg.Text('Quantidade:',font=('Heveltica',10,'bold')),sg.Input(size=(5),key='-quant')],
        [sg.Button('Adicionar Produto',key='-adicionar')]
    ]
    
    coluna2=[
        [sg.Text('Rascunho da Nota', font=('Haveltica',12,'bold'))],
        [sg.Button('Remover Produto'), sg.Button('Concluir Nota')],
        [sg.Listbox(values=nota,key='-note',size=(30,20))]
    ]
    
    coluna3=[
         [sg.Text('Lista de Produtos', font=('Haveltica',14,'bold'))],
         [sg.Button('Selecionar',key='-this')],
         [sg.Listbox(values=produtos,key='-prod',size=(22,20))]
    ]
    interface=[
        [sg.Text('Nova Nota', font='Haveltica 14')],
        [sg.Column(coluna3),sg.VerticalSeparator(),sg.Column(coluna1,size=(400,200)),sg.VerticalSeparator(),sg.Column(coluna2)]
        
    ]
    
    window=sg.Window('Criar nota').Layout(interface)
    
    stop=False
    while not stop:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            stop=True
        elif event == '-this':
            if len(values['-prod'])>0:
                name=values['-prod'][0]
                window.find_element('-nome').update(name)
                for pr in bd:
                    if pr[0]==name:
                        preco=pr[2]
                window.find_element('-preço').update(preco)
            else:
                janelaErro('Primeiro selecione um produto')
        elif event == '-adicionar':
            if len(values['-prod'])>0:
                if values['-quant'] != '':
                    final=[]
                    name=values['-prod'][0]
                    for pr in bd:
                        if pr[0]==name:
                            preco=pr[2]
                            if ',' in preco:
                                preçoNovo=preco.replace(',','.')
                                preço=float(preçoNovo)
                            else:
                                preço=float(preco)
                    quantidade=values['-quant']
                    final.append(name)
                    final.append(quantidade)
                    final.append('')
                    final.append('')
                    final.append(preço)
                    final.append('')
                    bd2.append(final)
                    bd2.sort(key=lambda x:x[0].upper())
                    nota.append(name +', '+quantidade)
                    nota.sort(key=lambda x:x[0].upper())
                    window.find_element('-note').update(nota)
                else:
                    janelaErro('Insira a quantidade do produto')
                    
            else:
                janelaErro('Primeiro selecione um produto')
        elif event == 'Remover Produto':
            if len(values['-note'])>0:
                name=values['-note'][0]
                novoNome=name.split(',')[-2]
                for elem in bd2:
                    if elem[0]==novoNome:
                        bd2.remove(elem)
                        bd2.sort(key=lambda x:x[0].upper())
                for prod in nota:
                    if novoNome in prod:
                        nota.remove(prod)
                nota.sort(key=lambda x:x[0].upper())
                window.find_element('-note').update(nota)
            else:
                janelaErro('Primeiro selecione um produto que se encontra no rascunho da nota')
        elif event == 'Concluir Nota':
            if len(bd2)>0:
                path1='C:/Users/rafay/Desktop/project/DONT_OPEN/'
                path2='C:/Users/rafay/OneDrive/Área de Trabalho/taty/Notas/'
                ficheiros=[f for f in listdir(path1) if isfile(join(path1,f))]
                ficheiros2=[f for f in listdir(path2) if isfile(join(path2,f))]
                fnome1='Nota '+str(len(ficheiros))+'.csv'
                fnome2='Nota '+str(len(ficheiros2)+1)+'.xlsx'
                file2=path2+fnome2
                file1=path1+fnome1
                md.criarNote(bd2,fnome1,path1)
                md.criarNotaExcel(file1,file2)
                wb=load_workbook(file2)
                ws=wb.active
                my_fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')
                my_font=Font(color='FFFFFF',bold=True)
                fonte2=Font(bold=True)
                my_fill2=PatternFill(start_color='EEECE1',end_color='EEECE1',fill_type='solid')
                alignment=Alignment(horizontal='center',vertical='center')
                my_header=['A1','B1','C1','D1','E1','F1']
                for cell in my_header:
                    ws[cell].fill=my_fill
                    ws[cell].font=my_font
                    ws[cell].alignment=alignment
                ws.column_dimensions['A'].width=46
                ws.column_dimensions['B'].width=11
                ws.column_dimensions['C'].width=8
                ws.column_dimensions['D'].width=8
                ws.column_dimensions['F'].width=13
                max_row = ws.max_row
                max_column=ws.max_column
                coluna_preco=[]
                for i in range(1, max_row + 1):
                    cell_obj = ws.cell(row = i, column = 5)
                    elemento=str(cell_obj)
                    novoElemento=elemento.split('.')[-1]
                    trueElemento=novoElemento.split('>')[-2]
                    coluna_preco.append(trueElemento)
                for elem in coluna_preco[1:]:
                    ws[elem].font=fonte2
                    ws[elem].number_format = '#,##0.00€' 
                thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
                tabela=[]
                for j in range (1,max_row+1):
                    for h in range (1,max_column+1):
                        celulaa=ws.cell(row=j,column=h)
                        elementoo=str(celulaa)
                        novoElementoo=elementoo.split('.')[-1]
                        trueElementoo=novoElementoo.split('>')[-2]
                        tabela.append(trueElementoo)
                for elemento in tabela:
                    ws[elemento].border=thin_border
                for elemen in tabela[6:]:
                    ws[elemen].fill=my_fill2
                wb.save(file2)
                window.close()
            else:
                janelaErro('Primeiro insira os produtos')
    window.close()

BD=md.lerDataset('C:/Users/rafay/Desktop/project/DONT_OPEN/produtos.csv')
sg.theme('Reddit')
janelaPrincipal=[
    [sg.Button('Adicionar Produto',key='-add',font=('Haveltica', 14)),sg.Button('Consultar Produto',key='-search',font=('Haveltica', 14)),sg.Button('Criar Nota', key='-note',font=('Haveltica', 14))]
]

window = sg.Window(title='Taty', font='Helvetica 24').Layout(janelaPrincipal)
stop = False 
while not stop:
    event, values = window.read()
    if event == sg.WIN_CLOSED: 
        stop = True
    elif event == '-search':
        consultarProduto(BD)
        BD.sort(key=lambda x:x[0].upper())
    elif event == '-add':
        novoProduto=janelaAdd(BD)
        if novoProduto != None:
            BD.append(novoProduto)
            BD.sort(key=lambda x:x[0].upper())
    elif event == '-note':
        criarNota(BD)
md.guardarBD(BD)
window.close()